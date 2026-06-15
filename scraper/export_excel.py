"""
Export buildings.geojson to an Excel workbook.
Produces: data/Burke_Heritage_Buildings.xlsx
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT      = Path(__file__).resolve().parent.parent
GEOJSON   = ROOT / "data" / "buildings.geojson"
OUT_XLSX  = ROOT / "data" / "Burke_Heritage_Buildings.xlsx"

COLUMNS = [
    ("Name / Building",       "name",             30),
    ("Alternate Name",        "alternate_name",   24),
    ("Address",               "address",          32),
    ("Neighbourhood",         "neighbourhood",    20),
    ("City",                  "city",             12),
    ("Year Built",            "year_built",       12),
    ("Architect / Firm",      "firm",             32),
    ("Architectural Style",   "style",            22),
    ("Building Type",         "building_type",    18),
    ("Current Use",           "current_use",      18),
    ("Former Use",            "former_use",       18),
    ("Heritage Status",       "heritage_status",  20),
    ("Heritage District",     "heritage_district",26),
    ("Awards",                "awards",           40),
    ("Notes",                 "notes",            50),
    ("Ward",                  "ward",             8),
    ("Source",                "source",           18),
    ("Latitude",              "_lat",             14),
    ("Longitude",             "_lng",             14),
    ("Thumbnail",             "thumbnail_url",    50),
    ("Wikipedia",             "wikipedia_url",    50),
    ("ACO / Heritage Detail", "detail_url",       50),
]

# ── Firm → colour (matches the map legend) ───────────────────────────────────
FIRM_FILLS = {
    "langley": PatternFill("solid", fgColor="BBDEFB"),   # light blue
    "solo":    PatternFill("solid", fgColor="FFE0B2"),   # light orange
    "horwood": PatternFill("solid", fgColor="C8E6C9"),   # light green
    "white":   PatternFill("solid", fgColor="E1BEE7"),   # light purple
    "other":   PatternFill("solid", fgColor="ECEFF1"),   # light grey
}

def classify(firm: str) -> str:
    f = firm.upper()
    if "WHITE"   in f: return "white"
    if "HORWOOD" in f: return "horwood"
    if "LANGLEY" in f: return "langley"
    if "BURKE"   in f: return "solo"
    return "other"

HEADER_FILL   = PatternFill("solid", fgColor="1A237E")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
THIN          = Side(style="thin", color="BDBDBD")
THIN_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

with open(GEOJSON, encoding="utf-8") as f:
    geojson = json.load(f)

features = geojson.get("features", [])

wb = Workbook()
ws = wb.active
ws.title = "Heritage Buildings"

# ── Header row ───────────────────────────────────────────────────────────────
for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font   = HEADER_FONT
    cell.fill   = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# ── Data rows ─────────────────────────────────────────────────────────────────
for row_idx, feat in enumerate(features, 2):
    p    = feat["properties"]
    coords = feat["geometry"]["coordinates"]  # [lng, lat]
    lat  = round(coords[1], 6)
    lng  = round(coords[0], 6)

    row_fill = FIRM_FILLS[classify(p.get("firm", ""))]

    for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
        if key == "_lat":
            value = lat
        elif key == "_lng":
            value = lng
        else:
            value = p.get(key, "") or ""

        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = BODY_FONT
        cell.fill      = row_fill
        cell.border    = THIN_BORDER
        wrap = key in ("notes", "awards")
        cell.alignment = Alignment(vertical="top", wrap_text=wrap)

        # Make URL columns clickable hyperlinks
        if key in ("wikipedia_url", "detail_url", "thumbnail_url") and value:
            cell.hyperlink = value
            cell.font = Font(name="Calibri", size=10, color="1565C0", underline="single")

ws.auto_filter.ref = ws.dimensions

# ── Legend sheet ─────────────────────────────────────────────────────────────
leg = wb.create_sheet("Legend")
leg.column_dimensions["A"].width = 28
leg.column_dimensions["B"].width = 40

leg_header_data = [
    ("Colour coding by firm / partnership", ""),
    ("Langley & Burke (1873–1892)",  "Light blue"),
    ("Edmund Burke — solo",          "Light orange"),
    ("Burke & Horwood",              "Light green"),
    ("Burke, Horwood & White",       "Light purple"),
    ("Other / Unknown",              "Light grey"),
]
for r, (label, note) in enumerate(leg_header_data, 1):
    ca = leg.cell(row=r, column=1, value=label)
    cb = leg.cell(row=r, column=2, value=note)
    if r == 1:
        ca.font = Font(bold=True, name="Calibri", size=11)
    else:
        key = list(FIRM_FILLS.keys())[r - 2]
        ca.fill = FIRM_FILLS[key]
        cb.fill = FIRM_FILLS[key]
        ca.font = BODY_FONT
        cb.font = BODY_FONT
    ca.border = THIN_BORDER
    cb.border = THIN_BORDER

wb.save(OUT_XLSX)
print(f"Wrote {len(features)} rows to {OUT_XLSX}")
